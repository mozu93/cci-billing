# app/ui/issuance_from_project.py
from datetime import date
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QHeaderView, QComboBox, QLineEdit, QMessageBox,
    QSpinBox, QFileDialog, QProgressDialog
)
from PyQt6.QtCore import Qt, QTimer
from app.database.connection import get_session
from app.services.project_service import (
    get_projects, get_project_members, get_project_templates
)
from app.services.category_service import get_active_categories
from app.services.issuance_service import create_issuance_for_member, mark_as_issued
from app.utils import current_user


COL_CHK  = 0
COL_NUM  = 1   # 会員番号
COL_ORG  = 2   # 事業所名
COL_KANA = 3   # フリガナ
COL_REP  = 4   # 代表者名
# 数量列: 5 〜 5+len(templates)-1
# 請求書列 = 5+len(templates)、領収書列 = 6+len(templates) — テンプレート数で可変


class _QtySpinBox(QSpinBox):
    """テーブル内数量入力用: Enter で同列の次行へフォーカスを移動する。"""

    def __init__(self, table: "QTableWidget", row: int, col: int):
        super().__init__()
        self._tbl = table
        self._row = row
        self._col = col

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            next_row = self._row + 1
            if next_row < self._tbl.rowCount():
                nxt = self._tbl.cellWidget(next_row, self._col)
                if nxt:
                    nxt.setFocus()
                    nxt.selectAll()
        else:
            super().keyPressEvent(event)


class _CheckableTable(QTableWidget):
    """チェックボックス列の Shift+クリック範囲選択に対応したテーブル。"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._last_checked_row: int = -1

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            idx = self.indexAt(event.pos())
            if idx.isValid() and idx.column() == COL_CHK:
                item = self.item(idx.row(), COL_CHK)
                if item and (item.flags() & Qt.ItemFlag.ItemIsUserCheckable):
                    if (event.modifiers() & Qt.KeyboardModifier.ShiftModifier
                            and self._last_checked_row >= 0):
                        new_state = (Qt.CheckState.Unchecked
                                     if item.checkState() == Qt.CheckState.Checked
                                     else Qt.CheckState.Checked)
                        r1 = min(self._last_checked_row, idx.row())
                        r2 = max(self._last_checked_row, idx.row())
                        self.blockSignals(True)
                        for r in range(r1, r2 + 1):
                            it = self.item(r, COL_CHK)
                            if it:
                                it.setCheckState(new_state)
                        self.blockSignals(False)
                        self._last_checked_row = idx.row()
                        return
                    else:
                        self._last_checked_row = idx.row()
        super().mousePressEvent(event)


class IssuanceFromProjectWidget(QWidget):
    def __init__(self, doc_type: str = "invoice"):
        super().__init__()
        self._doc_type = doc_type
        self._templates: list[dict] = []  # [{id, name}, ...]
        self._sort_col: int = COL_ORG
        self._sort_asc: bool = True
        self._qty_cache: dict[int, dict[int, int]] = {}  # {pm_id: {tmpl_id: qty}}
        self._all_projects: list = []
        self._build()
        self._load_projects()

    @property
    def _col_inv(self) -> int:
        return 5 + len(self._templates)

    @property
    def _col_rcp(self) -> int:
        return 6 + len(self._templates)

    def _build(self):
        layout = QVBoxLayout(self)

        top = QHBoxLayout()
        top.addWidget(QLabel("年度："))
        self._year_combo = QComboBox()
        self._year_combo.setMinimumWidth(95)
        self._year_combo.currentIndexChanged.connect(self._filter_projects)
        top.addWidget(self._year_combo)
        top.addWidget(QLabel("業務区分："))
        self._cat_combo = QComboBox()
        self._cat_combo.setMinimumWidth(120)
        self._cat_combo.currentIndexChanged.connect(self._filter_projects)
        top.addWidget(self._cat_combo)
        top.addWidget(QLabel("件名："))
        self._proj_combo = QComboBox()
        self._proj_combo.setMinimumWidth(240)
        self._proj_combo.currentIndexChanged.connect(self._on_project_changed)
        top.addWidget(self._proj_combo)
        top.addWidget(QLabel("表示："))
        self._filter_combo = QComboBox()
        self._filter_combo.addItems(["未発行のみ", "すべて"])
        self._filter_combo.currentIndexChanged.connect(self._load_members)
        top.addWidget(self._filter_combo)
        top.addStretch()
        layout.addLayout(top)

        search_row = QHBoxLayout()
        self._search = QLineEdit()
        self._search.setPlaceholderText("事業所名・代表者名で絞り込み")
        self._timer = QTimer()
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self._load_members)
        self._search.textChanged.connect(lambda: self._timer.start(300))
        search_row.addWidget(QLabel("検索："))
        search_row.addWidget(self._search)
        layout.addLayout(search_row)

        # ── ボタン行（テーブルの上） ──────────────────────────────────
        label = "請求書" if self._doc_type == "invoice" else "領収書"
        btn_row = QHBoxLayout()
        self._btn_issue = QPushButton(f"選択行に{label}を発行")
        self._btn_issue.clicked.connect(self._issue_checked)
        self._btn_issue_all = QPushButton(f"全員に{label}を発行")
        self._btn_issue_all.clicked.connect(self._issue_all)
        self._delivery_combo = QComboBox()
        self._delivery_combo.addItems(["窓口手渡し", "郵送", "メール送付", "その他"])
        btn_row.addWidget(self._btn_issue)
        btn_row.addWidget(self._btn_issue_all)
        btn_row.addWidget(QLabel("配付方法："))
        btn_row.addWidget(self._delivery_combo)
        btn_row.addSpacing(16)
        self._btn_export_xlsx = QPushButton("Excel出力")
        self._btn_export_xlsx.setToolTip(
            "表示中の名簿と数量をExcelに出力します。\n"
            "Excelで数量を入力後、「Excel取込」で読み込めます。")
        self._btn_export_xlsx.clicked.connect(self._export_excel)
        btn_row.addWidget(self._btn_export_xlsx)
        self._btn_import_xlsx = QPushButton("Excel取込")
        self._btn_import_xlsx.setToolTip(
            "Excel出力したファイルを読み込み、数量と発行チェックを画面に反映します。")
        self._btn_import_xlsx.clicked.connect(self._import_excel)
        btn_row.addWidget(self._btn_import_xlsx)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self._table = _CheckableTable(0, 7)
        hdr_item = QTableWidgetItem()
        hdr_item.setCheckState(Qt.CheckState.Unchecked)
        self._table.setHorizontalHeaderItem(COL_CHK, hdr_item)
        self._table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self._table.horizontalHeader().setSortIndicatorShown(True)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._table.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._setup_table_columns()
        layout.addWidget(self._table)

        self._status_label = QLabel("")
        layout.addWidget(self._status_label)

    def _setup_table_columns(self):
        n = len(self._templates)
        self._table.setColumnCount(7 + n)
        headers = ["", "会員番号", "事業所名", "フリガナ", "代表者名"]
        for tmpl in self._templates:
            headers.append(f"{tmpl['name']}\n数量")
        headers += ["請求書", "領収書"]
        self._table.setHorizontalHeaderLabels(headers)

        hdr = self._table.horizontalHeader()
        fixed = QHeaderView.ResizeMode.Fixed
        interactive = QHeaderView.ResizeMode.Interactive
        hdr.setSectionResizeMode(COL_CHK,  fixed);      self._table.setColumnWidth(COL_CHK,  30)
        hdr.setSectionResizeMode(COL_NUM,  interactive); self._table.setColumnWidth(COL_NUM,  80)
        hdr.setSectionResizeMode(COL_ORG,  interactive); self._table.setColumnWidth(COL_ORG, 180)
        hdr.setSectionResizeMode(COL_KANA, interactive); self._table.setColumnWidth(COL_KANA,140)
        hdr.setSectionResizeMode(COL_REP,  interactive); self._table.setColumnWidth(COL_REP, 100)
        for col in range(5, 5 + n):
            hdr.setSectionResizeMode(col, interactive)
            self._table.setColumnWidth(col, 120)
        for col in (self._col_inv, self._col_rcp):
            hdr.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)

        hdr.setSortIndicator(
            self._sort_col,
            Qt.SortOrder.AscendingOrder if self._sort_asc else Qt.SortOrder.DescendingOrder)

    # ── ヘッダークリック：全選択 / 全解除 ─────────────────────────

    def _on_header_clicked(self, col: int):
        if col == COL_CHK:
            hdr = self._table.horizontalHeaderItem(COL_CHK)
            if hdr is None:
                return
            new_state = (Qt.CheckState.Unchecked
                         if hdr.checkState() == Qt.CheckState.Checked
                         else Qt.CheckState.Checked)
            hdr.setCheckState(new_state)
            self._table.blockSignals(True)
            for r in range(self._table.rowCount()):
                it = self._table.item(r, COL_CHK)
                if it:
                    it.setCheckState(new_state)
            self._table.blockSignals(False)
        else:
            # 数量 SpinBox 列はソート対象外
            qty_cols = set(range(5, 5 + len(self._templates)))
            if col in qty_cols:
                return
            self._save_qty_cache()
            if self._sort_col == col:
                self._sort_asc = not self._sort_asc
            else:
                self._sort_col = col
                self._sort_asc = True
            self._load_members()

    # ── プロジェクト読み込み ───────────────────────────────────────

    def showEvent(self, event):
        super().showEvent(event)
        self._load_projects()

    def _load_projects(self):
        session = get_session()
        try:
            self._all_projects = get_projects(session, status="active")
            cats = get_active_categories(session)
        finally:
            session.close()

        # 年度コンボ（重複なし降順）
        years = sorted({p.fiscal_year for p in self._all_projects}, reverse=True)
        current_year = self._year_combo.currentData()
        self._year_combo.blockSignals(True)
        self._year_combo.clear()
        self._year_combo.addItem("すべて", None)
        for y in years:
            self._year_combo.addItem(f"{y}年度", y)
        # デフォルト：最新年度
        if years and current_year is None:
            self._year_combo.setCurrentIndex(1)
        else:
            for i in range(self._year_combo.count()):
                if self._year_combo.itemData(i) == current_year:
                    self._year_combo.setCurrentIndex(i)
                    break
        self._year_combo.blockSignals(False)

        # 業務区分コンボ：_all_projects に含まれるカテゴリのみ表示（窓口除外済み）
        used_cat_ids = {p.category_id for p in self._all_projects}
        current_cat = self._cat_combo.currentData()
        self._cat_combo.blockSignals(True)
        self._cat_combo.clear()
        self._cat_combo.addItem("すべて", None)
        for c in cats:
            if c.id in used_cat_ids:
                self._cat_combo.addItem(c.name, c.id)
        for i in range(self._cat_combo.count()):
            if self._cat_combo.itemData(i) == current_cat:
                self._cat_combo.setCurrentIndex(i)
                break
        self._cat_combo.blockSignals(False)

        self._filter_projects()

    def _filter_projects(self):
        sel_year = self._year_combo.currentData()
        sel_cat  = self._cat_combo.currentData()
        current_id = self._proj_combo.currentData()
        self._proj_combo.blockSignals(True)
        self._proj_combo.clear()
        for p in self._all_projects:
            if sel_year is not None and p.fiscal_year != sel_year:
                continue
            if sel_cat is not None and p.category_id != sel_cat:
                continue
            self._proj_combo.addItem(p.name, p.id)
        for i in range(self._proj_combo.count()):
            if self._proj_combo.itemData(i) == current_id:
                self._proj_combo.setCurrentIndex(i)
                break
        self._proj_combo.blockSignals(False)
        self._on_project_changed()

    def _on_project_changed(self):
        project_id = self._proj_combo.currentData()
        if project_id is None:
            self._templates = []
        else:
            session = get_session()
            try:
                pts = get_project_templates(session, project_id)
                self._templates = [
                    {"id": pt.item_template.id, "name": pt.item_template.name}
                    for pt in pts
                ]
            finally:
                session.close()
        self._setup_table_columns()
        self._load_members()

    # ── メンバー一覧読み込み ──────────────────────────────────────

    _STATUS_SHORT = {"発行済み": "発行済", "支払済み": "支払済", "準備中": "準備中"}

    def _cell_text(self, iss) -> str:
        if iss is None:
            return "未発行"
        short = self._STATUS_SHORT.get(iss.status, iss.status)
        return f"{short} {iss.doc_number}".strip()

    def _load_members(self):
        project_id = self._proj_combo.currentData()
        if project_id is None:
            self._table.setRowCount(0)
            return
        query = self._search.text().strip().lower()
        show_all = self._filter_combo.currentIndex() == 1
        doc_type = self._doc_type
        session = get_session()
        try:
            pms = get_project_members(session, project_id)
            from app.database.models import Issuance
            pm_data = []
            for pm in pms:
                inv = (session.query(Issuance)
                       .filter_by(project_member_id=pm.id, doc_type="invoice")
                       .order_by(Issuance.created_at.desc())
                       .first())
                rcp = (session.query(Issuance)
                       .filter_by(project_member_id=pm.id, doc_type="receipt")
                       .order_by(Issuance.created_at.desc())
                       .first())
                voided = inv is None and rcp is not None
                sel = inv if doc_type == "invoice" else rcp
                sel_status = sel.status if sel else "未発行"
                hide_issued = sel_status in ("発行済み", "支払済み")
                hide_voided = doc_type == "invoice" and voided
                if not show_all and (hide_issued or hide_voided):
                    continue
                if query:
                    targets = [
                        pm.organization_name or "",
                        pm.representative_name or "",
                        pm.organization_kana or "",
                    ]
                    if not any(query in t.lower() for t in targets):
                        continue
                inv_text = "無効" if voided else self._cell_text(inv)
                pm_data.append((
                    pm.id, pm,
                    inv_text, self._cell_text(rcp),
                    inv.id if inv else None, rcp.id if rcp else None,
                ))
        finally:
            session.close()

        # ソート
        col_inv = self._col_inv
        col_rcp = self._col_rcp
        sc = self._sort_col

        def _key(item):
            _, pm, inv_text, rcp_text, _, _ = item
            if sc == COL_NUM:   return pm.member_number or ""
            if sc == COL_ORG:   return pm.organization_name or ""
            if sc == COL_KANA:  return pm.organization_kana or ""
            if sc == COL_REP:   return pm.representative_name or ""
            if sc == col_inv:   return inv_text
            if sc == col_rcp:   return rcp_text
            return ""

        pm_data.sort(key=_key, reverse=not self._sort_asc)

        self._table._last_checked_row = -1
        hdr_item = self._table.horizontalHeaderItem(COL_CHK)
        if hdr_item:
            hdr_item.setCheckState(Qt.CheckState.Unchecked)

        self._table.setRowCount(0)
        for pm_id, pm, inv_text, rcp_text, inv_id, rcp_id in pm_data:
            row = self._table.rowCount()
            self._table.insertRow(row)

            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsUserCheckable)
            chk_item.setCheckState(Qt.CheckState.Unchecked)
            self._table.setItem(row, COL_CHK, chk_item)

            row_data = (pm_id, inv_id, rcp_id)
            for col, val in [
                (COL_NUM,  pm.member_number or ""),
                (COL_ORG,  pm.organization_name or ""),
                (COL_KANA, pm.organization_kana or ""),
                (COL_REP,  pm.representative_name or ""),
                (col_inv,  inv_text),
                (col_rcp,  rcp_text),
            ]:
                it = QTableWidgetItem(val)
                it.setData(Qt.ItemDataRole.UserRole, row_data)
                self._table.setItem(row, col, it)

            for col_offset, tmpl in enumerate(self._templates):
                col = 5 + col_offset
                spin = _QtySpinBox(self._table, row, col)
                spin.setRange(0, 9999)
                cached = self._qty_cache.get(pm_id, {}).get(tmpl["id"], 1)
                spin.setValue(cached)
                spin.setAlignment(Qt.AlignmentFlag.AlignCenter)
                spin.setStyleSheet(
                    "QSpinBox { min-height: 0; padding: 1px 4px; }")
                self._table.setCellWidget(row, col, spin)

        self._table.horizontalHeader().setSortIndicator(
            self._sort_col,
            Qt.SortOrder.AscendingOrder if self._sort_asc else Qt.SortOrder.DescendingOrder)
        self._table.resizeRowsToContents()
        self._status_label.setText(f"{len(pm_data)} 件表示")

    # ── 行数量取得 / キャッシュ保存 ──────────────────────────────

    def _get_row_quantities(self, row: int) -> dict[int, int]:
        result = {}
        for col_offset, tmpl in enumerate(self._templates):
            spin = self._table.cellWidget(row, 5 + col_offset)
            if isinstance(spin, _QtySpinBox):
                result[tmpl["id"]] = spin.value()
        return result

    def _save_qty_cache(self):
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if not data_item:
                continue
            pm_id, _, _ = data_item.data(Qt.ItemDataRole.UserRole)
            self._qty_cache[pm_id] = self._get_row_quantities(r)

    # ── Excel入出力 ───────────────────────────────────────────────

    _XLSX_FIXED_HEADERS = ["ID", "会員番号", "事業所名", "フリガナ", "代表者名"]

    def _export_excel(self):
        """表示中の名簿＋項目ごとの数量をExcelに出力する。"""
        project_id = self._proj_combo.currentData()
        if project_id is None:
            QMessageBox.information(self, "案件未選択", "件名を選択してください。")
            return
        if not self._templates:
            QMessageBox.information(
                self, "項目なし",
                "この案件には項目テンプレートが設定されていません。")
            return
        if self._table.rowCount() == 0:
            QMessageBox.information(self, "対象なし", "表示中の名簿がありません。")
            return

        proj_name = self._proj_combo.currentText()
        safe = "".join(c for c in proj_name if c not in '\\/:*?"<>|')
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel出力", f"{safe}_数量入力.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数量入力"
        headers = self._XLSX_FIXED_HEADERS + [t["name"] for t in self._templates]
        ws.append(headers)
        fill = PatternFill("solid", fgColor="DDEBF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill

        exported = 0
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if not data_item:
                continue
            pm_id, _, _ = data_item.data(Qt.ItemDataRole.UserRole)
            qty = self._get_row_quantities(r)
            ws.append([
                pm_id,
                self._table.item(r, COL_NUM).text(),
                self._table.item(r, COL_ORG).text(),
                self._table.item(r, COL_KANA).text(),
                self._table.item(r, COL_REP).text(),
            ] + [qty.get(t["id"], 0) for t in self._templates])
            exported += 1

        widths = [6, 10, 28, 22, 14] + [14] * len(self._templates)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.critical(
                self, "保存エラー",
                "ファイルを保存できませんでした。\n"
                "同じファイルをExcelで開いたままになっていないか確認してください。")
            return
        QMessageBox.information(
            self, "Excel出力",
            f"{exported}件を出力しました。\n{path}\n\n"
            "Excelで数量を入力・保存後、「Excel取込」で読み込んでください。\n"
            "・数量0の項目は明細に含まれません\n"
            "・全項目0の行は発行対象外になります\n"
            "・ID列は照合に使うため変更しないでください")

    def _import_excel(self):
        """Excel出力で編集したファイルを読み込み、数量とチェックを画面に反映する。"""
        if self._table.rowCount() == 0:
            QMessageBox.information(self, "対象なし", "表示中の名簿がありません。")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel取込", "", "Excel (*.xlsx)")
        if not path:
            return

        import openpyxl
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "読込エラー", str(e))
            return
        if not all_rows:
            QMessageBox.warning(self, "読込エラー", "データが見つかりませんでした。")
            return

        header = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        if "ID" not in header:
            QMessageBox.critical(
                self, "読込エラー",
                "見出し行に「ID」列が見つかりません。\n"
                "「Excel出力」で出力したファイルを使用してください。")
            return
        id_col = header.index("ID")
        tmpl_cols: dict[int, int] = {}
        missing_cols: list[str] = []
        for t in self._templates:
            if t["name"] in header:
                tmpl_cols[t["id"]] = header.index(t["name"])
            else:
                missing_cols.append(t["name"])
        if not tmpl_cols:
            QMessageBox.critical(
                self, "読込エラー",
                "この案件の項目に対応する数量列が1つも見つかりません。\n"
                "案件の選択が出力時と同じか確認してください。")
            return

        file_qty: dict[int, dict[int, int]] = {}
        bad_rows: list[str] = []
        for i, cells in enumerate(all_rows[1:], start=2):
            raw_id = cells[id_col] if id_col < len(cells) else None
            if raw_id is None or str(raw_id).strip() == "":
                continue
            try:
                pm_id = int(str(raw_id).strip())
            except ValueError:
                bad_rows.append(f"{i}行目: ID「{raw_id}」が数値ではありません")
                continue
            q = {}
            for tid, col in tmpl_cols.items():
                v = cells[col] if col < len(cells) else None
                if v is None or str(v).strip() == "":
                    q[tid] = 0
                    continue
                try:
                    f = float(str(v).strip())
                    iv = int(f)
                    if iv < 0:
                        raise ValueError
                    if iv != f:
                        bad_rows.append(
                            f"{i}行目: 数量「{v}」は整数でないため{iv}として扱います")
                    if iv > 9999:
                        bad_rows.append(
                            f"{i}行目: 数量「{v}」は上限の9999に丸めました")
                        iv = 9999
                    q[tid] = iv
                except (ValueError, OverflowError):
                    bad_rows.append(f"{i}行目: 数量「{v}」が不正です（0として扱います）")
                    q[tid] = 0
            file_qty[pm_id] = q

        applied = 0
        checked = 0
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if not data_item:
                continue
            pm_id, _, _ = data_item.data(Qt.ItemDataRole.UserRole)
            if pm_id not in file_qty:
                continue
            q = file_qty.pop(pm_id)
            for col_offset, tmpl in enumerate(self._templates):
                spin = self._table.cellWidget(r, 5 + col_offset)
                if isinstance(spin, _QtySpinBox) and tmpl["id"] in q:
                    spin.setValue(q[tmpl["id"]])
            total = sum(q.values())
            chk = self._table.item(r, COL_CHK)
            if chk:
                chk.setCheckState(Qt.CheckState.Checked if total > 0
                                  else Qt.CheckState.Unchecked)
            if total > 0:
                checked += 1
            applied += 1
        self._save_qty_cache()

        msg = [f"{applied}件の数量を反映し、{checked}件に発行チェックを入れました。",
               "内容を確認のうえ「選択行に発行」ボタンで発行してください。"]
        if file_qty:
            msg.append(f"※名簿に表示されていないID {len(file_qty)}件は反映できませんでした。\n"
                       "（発行済み等で非表示の可能性。表示を「すべて」にして再度取り込んでください）")
        if missing_cols:
            msg.append("※次の項目の数量列が見つかりませんでした：" + "、".join(missing_cols))
        if bad_rows:
            shown = "\n".join(bad_rows[:5])
            more = f"\n…ほか{len(bad_rows) - 5}件" if len(bad_rows) > 5 else ""
            msg.append(f"※読み込めなかった値：\n{shown}{more}")
        QMessageBox.information(self, "Excel取込", "\n\n".join(msg))

    # ── チェック済み行の取得 ──────────────────────────────────────

    def _checked_rows(self) -> list[tuple[int, tuple]]:
        result = []
        for r in range(self._table.rowCount()):
            chk = self._table.item(r, COL_CHK)
            if chk and chk.checkState() == Qt.CheckState.Checked:
                data_item = self._table.item(r, COL_ORG)
                if data_item:
                    result.append((r, data_item.data(Qt.ItemDataRole.UserRole)))
        return result

    # ── 発行処理 ──────────────────────────────────────────────────

    def _do_issue_rows(self, rows: list[tuple[int, tuple]]) -> list[str]:
        """rows = [(row_idx, (pm_id, inv_id, rcp_id)), ...] を発行して PDF 生成。
        エラーメッセージのリストを返す。"""
        project_id = self._proj_combo.currentData()
        doc_type = self._doc_type
        delivery = self._delivery_combo.currentText()
        session = get_session()
        errors = []
        issued_issuances = []
        try:
            from app.database.models import ProjectMember, Issuance
            from app.utils.pdf_helpers import generate_and_open
            for row_idx, (pm_id, invoice_id, receipt_id) in rows:
                if doc_type == "invoice" and invoice_id is None and receipt_id is not None:
                    continue
                quantities = self._get_row_quantities(row_idx)
                issuance_id = invoice_id if doc_type == "invoice" else receipt_id
                if issuance_id is None and quantities and not any(quantities.values()):
                    org_item = self._table.item(row_idx, COL_ORG)
                    name = org_item.text() if org_item else f"{row_idx + 1}行目"
                    errors.append(f"{name}：数量がすべて0のためスキップしました")
                    continue
                try:
                    pm = session.get(ProjectMember, pm_id)
                    if issuance_id is None:
                        today = date.today()
                        iss = create_issuance_for_member(
                            session, project_id=project_id,
                            project_member_id=pm_id,
                            recipient_organization=pm.organization_name,
                            recipient_name=pm.representative_name,
                            doc_type=doc_type,
                            fiscal_year=today.year, month=today.month,
                            quantities=quantities if quantities else None,
                        )
                        issuance_id = iss.id

                    iss = session.get(Issuance, issuance_id)
                    if iss is None:
                        continue
                    if iss.status != "発行済み":
                        mark_as_issued(session, issuance_id,
                                       staff_id=current_user.get_id(),
                                       staff_name=current_user.get_name(),
                                       delivery_method=delivery)
                        iss = session.get(Issuance, issuance_id)
                    elif (iss.delivery_method or "") != delivery:
                        # 発行済み行の再実行時も配付方法を実態に合わせる
                        iss.delivery_method = delivery
                        session.commit()
                    issued_issuances.append((iss, session))
                except Exception as e:
                    errors.append(str(e))

            due_date = None
            if doc_type == "invoice" and issued_issuances:
                from app.ui.invoice_options_dialog import InvoiceOptionsDialog
                from PyQt6.QtWidgets import QDialog
                ref_iss = issued_issuances[0][0]
                opts = InvoiceOptionsDialog(issued_at=ref_iss.issued_at, parent=self)
                if opts.exec() != QDialog.DialogCode.Accepted:
                    return errors
                due_date = opts.due_date()

            for iss, sess in issued_issuances:
                try:
                    generate_and_open(iss, sess, due_date=due_date)
                except Exception as e:
                    errors.append(str(e))

            if delivery == "メール送付" and issued_issuances:
                self._send_issue_emails(issued_issuances, errors)
        finally:
            session.close()
        return errors

    def _send_issue_emails(self, issued_issuances: list, errors: list[str]):
        """配付方法「メール送付」で発行した分のPDFをメール添付で送信する。

        1つのSMTP接続を使い回し、進捗ダイアログ（中止可）を表示する。
        送信成否は操作ログに記録する。
        """
        from app.database.models import ProjectMember
        from app.services.email_service import SmtpSession, send_issuance_email
        from app.services.operation_log_service import add_log
        label = "請求書" if self._doc_type == "invoice" else "領収書"

        rows = []
        lines = []
        for iss, sess in issued_issuances:
            email = ""
            if iss.project_member_id:
                pm = sess.get(ProjectMember, iss.project_member_id)
                email = (pm.email or "").strip() if pm else ""
            rows.append((iss, sess, email))
            name = (iss.recipient_organization or iss.recipient_name
                    or iss.doc_number)
            lines.append(f"・{name} → {email or '（アドレス未登録）'}")

        shown = lines[:10]
        if len(lines) > 10:
            shown.append(f"…ほか{len(lines) - 10}件")
        box = QMessageBox(
            QMessageBox.Icon.Question, "メール送信の確認",
            f"{len(lines)}件の宛先に{label}をメール送信します。"
            f"よろしいですか？\n\n" + "\n".join(shown),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            self)
        box.setDefaultButton(QMessageBox.StandardButton.No)
        if box.exec() != QMessageBox.StandardButton.Yes:
            return

        progress = QProgressDialog(
            f"{label}をメール送信中…", "中止", 0, len(rows), self)
        progress.setWindowTitle("メール送信")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)

        sent = 0
        canceled_at = None
        try:
            with SmtpSession() as smtp:
                for i, (iss, sess, email) in enumerate(rows):
                    if progress.wasCanceled():
                        canceled_at = i
                        break
                    name = (iss.recipient_organization or iss.recipient_name
                            or iss.doc_number)
                    progress.setLabelText(
                        f"送信中（{i + 1}/{len(rows)}）：{name}")
                    progress.setValue(i)
                    try:
                        addr = send_issuance_email(
                            sess, iss, to_addr=email or None, smtp=smtp)
                        sent += 1
                        add_log(sess, "メール送信", "issuance", iss.id,
                                f"{label} {iss.doc_number} → {addr}")
                    except Exception as e:
                        errors.append(f"メール送信失敗：{e}")
                        add_log(sess, "メール送信失敗", "issuance", iss.id,
                                f"{label} {iss.doc_number}：{e}")
        except Exception as e:
            # SMTP接続自体の失敗（サーバー設定誤り等）
            errors.append(f"メール送信失敗（SMTP接続エラー）：{e}")
        finally:
            progress.setValue(len(rows))

        msg = f"{sent}件のメールを送信しました。"
        if canceled_at is not None:
            msg += f"\n（中止のため残り{len(rows) - canceled_at}件は未送信です）"
        QMessageBox.information(self, "メール送信", msg)

    def _issue_checked(self):
        targets = self._checked_rows()
        if not targets:
            QMessageBox.information(self, "未選択",
                                    "発行する行のチェックボックスにチェックを入れてください。")
            return
        errors = self._do_issue_rows(targets)
        if errors:
            QMessageBox.critical(self, "PDF生成エラー", "\n".join(errors))
        self._load_members()

    def _issue_all(self):
        project_id = self._proj_combo.currentData()
        if project_id is None:
            return
        label = "請求書" if self._doc_type == "invoice" else "領収書"
        ans = QMessageBox.question(
            self, "確認",
            f"表示中の全員ぶんを{label}で発行します。よろしいですか？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if ans != QMessageBox.StandardButton.Yes:
            return
        all_rows = []
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if data_item:
                all_rows.append((r, data_item.data(Qt.ItemDataRole.UserRole)))
        errors = self._do_issue_rows(all_rows)
        if errors:
            QMessageBox.critical(self, "PDF生成エラー", "\n".join(errors))
        self._load_members()
